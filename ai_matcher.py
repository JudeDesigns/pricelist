"""
AI-Powered Product Matching System
Uses Gemini AI to match extracted products against a reference database
"""

import os
from typing import List, Dict, Optional
from openpyxl import load_workbook
import google.generativeai as genai
import re
import json
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Configure Gemini API
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY')
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)


def load_reference_database(filepath: str = 'Book1.xlsx') -> List[Dict[str, str]]:
    """
    Load reference product database from Excel file.
    
    Args:
        filepath: Path to the Excel file (default: Book1.xlsx)
    
    Returns:
        List of dictionaries with keys: product_id, description
        Example: [
            {"product_id": "Z75880", "description": "Ground Beef 80/20"},
            {"product_id": "Z75881", "description": "Chicken Breast"},
            ...
        ]
    
    Raises:
        FileNotFoundError: If the Excel file doesn't exist
        ValueError: If the Excel file is malformed or missing required columns
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Reference database file not found: {filepath}")
    
    try:
        # Load the workbook
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        
        # Find header row and column indices
        product_name_col = None
        product_code_col = None
        
        # Check first row for headers
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                header = str(cell.value).strip().lower()
                if header in ['product name', 'product description', 'description', 'product_name']:
                    product_name_col = col_idx
                elif header in ['productcode', 'product code', 'product_code', 'product id', 'sku']:
                    product_code_col = col_idx
        
        if not product_name_col or not product_code_col:
            raise ValueError(
                f"Could not find required columns. Found headers: "
                f"{[cell.value for cell in ws[1]]}"
            )
        
        # Extract all products
        reference_products = []
        for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            product_code = ws.cell(row=row_idx, column=product_code_col).value
            product_name = ws.cell(row=row_idx, column=product_name_col).value

            if product_code and product_name:
                product_id_str = str(product_code).strip()

                # Skip products with -SP or SP suffix (special/promotional variants)
                if product_id_str.endswith('SP') or product_id_str.endswith('-SP'):
                    continue

                # Skip ZTBD products (they should match to K002XXX or K003XXX instead)
                if product_id_str.startswith('ZTBD'):
                    continue

                # Skip TEMPGLEN products (they should not be matched)
                if product_id_str.startswith('TEMPGLEN'):
                    continue

                reference_products.append({
                    'product_id': product_id_str,
                    'description': str(product_name).strip()
                })
        
        wb.close()
        
        if not reference_products:
            raise ValueError("No products found in reference database")
        
        return reference_products
        
    except Exception as e:
        raise ValueError(f"Error reading reference database: {str(e)}")


def normalize_product_id(product_id: str) -> str:
    """
    Normalize a product ID by extracting the core part.

    Rules:
    1. Strip ALL leading patterns (Z, K, G, G0, G00, K0, K00, etc.) before the numeric core
    2. Keep trailing suffixes with hyphen + 2 digits (e.g., -01, -02) - they're part of the ID
    3. Strip trailing letters that don't have a hyphen (they're not part of the ID)
    4. Special handling for K002XXX pattern: K002013-01 -> 13-01, K002014-01 -> 14-01

    Examples:
        Leading pattern stripping:
        "Z156171" -> "156171"
        "K156171" -> "156171"
        "G156171" -> "156171"
        "G015050194" -> "15050194"
        "G0015050194" -> "15050194"
        "K015050194" -> "15050194"

        Special K002XXX pattern:
        "K002013-01" -> "13-01"
        "K002014-01" -> "14-01"
        "K002019-01" -> "19-01"

        Trailing with hyphen + 2 digits (KEEP):
        "156171-01" -> "156171-01"
        "156171-02" -> "156171-02"
        "Z156171-01" -> "156171-01"
        "G015050194-01" -> "15050194-01"

        Trailing without hyphen (STRIP):
        "156171A" -> "156171"
        "156171B" -> "156171"
        "Z156171B" -> "156171"

    Args:
        product_id: Raw product ID string

    Returns:
        Normalized product ID
    """
    if not product_id or product_id == 'N/A':
        return ''

    # Clean and uppercase
    cleaned = str(product_id).strip().upper()

    # Special case: K002XXX-YY pattern (e.g., K002013-01, K002014-01)
    # Extract the last 2-3 digits before the hyphen as the core ID
    # Handles: K002013-01, K002013-01SP, K002013-01-SP, K002013-01FZ, etc.
    k002_match = re.match(r'^K002(\d{3})(-\d{2})(?:[-A-Z]*)$', cleaned)
    if k002_match:
        # Extract last 2-3 digits: 013 -> 13, 014 -> 14, 019 -> 19
        core_digits = k002_match.group(1).lstrip('0') or '0'
        suffix = k002_match.group(2)
        return core_digits + suffix

    # Strategy:
    # 1. Remove any prefix patterns (letters, leading zeros, hyphens, etc.) before the main number
    # 2. Extract: numeric core + optional hyphen + 2-digit suffix
    # 3. Strip trailing letters that don't have a hyphen
    # 4. Strip invalid suffixes (not exactly 2 digits)

    # Pattern explanation:
    # ^.*?        - Non-greedy match any prefix (TYS-, Z, K-, G0, G00, etc.)
    # (\d+)       - Capture the main numeric core (required)
    # (?:-\d+)?   - Non-capturing: optional hyphen + any digits (to skip over)
    # .*$         - Match rest of string

    # First, try to match with valid 2-digit suffix
    match = re.match(r'^.*?(\d+)(-\d{2})(?:[A-Z]*)$', cleaned)
    if match:
        # Return the numeric core + valid 2-digit suffix
        # Strip leading zeros from the numeric core
        numeric_core = match.group(1).lstrip('0') or '0'
        return numeric_core + match.group(2)

    # If no valid 2-digit suffix, extract just the main numeric core
    # This handles cases like "156171-1" or "156171-123" or "156171A"
    match = re.match(r'^.*?(\d+)', cleaned)
    if match:
        # Return just the numeric core, strip leading zeros
        numeric_core = match.group(1).lstrip('0') or '0'
        return numeric_core

    # Fallback: try to extract just the longest numeric sequence
    numeric_parts = re.findall(r'\d+', cleaned)
    if numeric_parts:
        longest = max(numeric_parts, key=len)
        return longest.lstrip('0') or '0'

    return cleaned


def select_simplest_product_id(matching_products: List[Dict[str, str]]) -> Dict[str, str]:
    """
    Select the simplest product ID from a list of matching products.
    Prefers IDs without extra suffixes or characters.

    Examples:
        K002019-01, K002019-01FZ, K002019-01FZ-SABYS -> K002019-01
        Kruse01, CG00001 -> Kruse01 (prefer Kruse over CG)
        Kruse05, Kruse5 -> Use description to disambiguate (Kruse5=bacon, Kruse05=ham)

    Returns:
        The product with the simplest ID, or None if no clear winner
    """
    if not matching_products:
        return None

    if len(matching_products) == 1:
        return matching_products[0]

    # Special case: Kruse01 vs CG00001 - always prefer Kruse01
    product_ids = [p['product_id'].upper() for p in matching_products]
    if 'KRUSE01' in product_ids and 'CG00001' in product_ids:
        for p in matching_products:
            if p['product_id'].upper() == 'KRUSE01':
                print(f"  ✅  Auto-selected Kruse01 over CG00001")
                return p

    # Special case: Prefer any Kruse product over CG products
    kruse_products = [p for p in matching_products if p['product_id'].upper().startswith('KRUSE')]
    cg_products = [p for p in matching_products if p['product_id'].upper().startswith('CG')]
    if kruse_products and cg_products and len(matching_products) == len(kruse_products) + len(cg_products):
        # Only Kruse and CG products - prefer Kruse
        if len(kruse_products) == 1:
            print(f"  ✅  Auto-selected {kruse_products[0]['product_id']} over CG product")
            return kruse_products[0]

    # Special case: Kruse5 vs Kruse05 - need to use description to disambiguate
    # Kruse5 = APPLEWOOD ROASTED SLICED BACON
    # Kruse05 = BONELESS BLACK FOREST HAM
    # Return None to let AI handle this based on description
    if 'KRUSE5' in product_ids and 'KRUSE05' in product_ids:
        return None  # Let AI disambiguate based on description

    # Score each product ID by complexity (lower is simpler)
    def complexity_score(product_id: str) -> tuple:
        """
        Return a tuple for sorting: (has_extra_suffix, is_not_kruse, length, alphabetical)
        Lower values = simpler
        """
        pid = str(product_id).upper()

        # Count extra suffixes after the base pattern
        # K002019-01FZ has extra suffix "FZ"
        # K002019-01FZ-SABYS has extra suffix "FZ-SABYS"
        has_extra_suffix = 0

        # Check for extra characters after standard patterns
        # Pattern: K002XXX-YY should not have anything after
        if re.match(r'^K002\d{3}-\d{2}[A-Z-]+', pid):
            has_extra_suffix = 1
        # Pattern: KruseXX should not have extra suffixes
        elif re.match(r'^KRUSE\d+[A-Z-]+', pid):
            has_extra_suffix = 1

        # Prefer Kruse products over CG products
        is_not_kruse = 0 if pid.startswith('KRUSE') else 1

        return (has_extra_suffix, is_not_kruse, len(pid), pid)

    # Sort by complexity
    sorted_products = sorted(matching_products, key=lambda p: complexity_score(p['product_id']))

    # Check if there's a clear winner (significantly simpler than others)
    simplest = sorted_products[0]
    simplest_score = complexity_score(simplest['product_id'])

    # If the simplest has no extra suffix and others do, it's a clear winner
    if simplest_score[0] == 0:  # No extra suffix
        # Check if any other product also has no extra suffix
        other_simple = [p for p in sorted_products[1:] if complexity_score(p['product_id'])[0] == 0]
        if not other_simple:
            # Only one product without extra suffix - clear winner
            return simplest

    # No clear winner
    return None


def match_product_by_id(
    extracted_product: Dict[str, str],
    reference_database: List[Dict[str, str]],
    confidence_threshold: float = 0.8,
    use_ai_fallback: bool = True
) -> Dict[str, any]:
    """
    Match an extracted product against reference database using ID-only matching.
    Falls back to AI matching if Product ID is missing/invalid or ambiguous.

    This function normalizes both extracted and reference IDs by:
    - Stripping leading patterns (Z, K, G, G0, G00, K0, K00, etc.)
    - Stripping trailing letters without hyphen
    - Keeping trailing hyphen + 2-digit suffixes

    Special handling for short IDs (1-2 digits):
    - If multiple products match the same short ID, uses AI to verify the correct match
    - This prevents incorrect matches like "06" matching "ZTBD6" when it should match "Kruse06"

    Args:
        extracted_product: Dict with keys 'product_id' and 'description'
        reference_database: List of reference products from load_reference_database()
        confidence_threshold: Minimum confidence for AI matching (default: 0.8)
        use_ai_fallback: If True, use AI matching when ID is missing or ambiguous (default: True)

    Returns:
        Dictionary with keys:
        - matched_id: The matched reference product ID or "NO_MATCH"
        - confidence: 1.0 for ID matches, 0.0-1.0 for AI matches
        - reasoning: Explanation of the match
        - matched_description: Description of matched product or None
    """
    # Handle empty reference database
    if not reference_database:
        return {
            'matched_id': 'NO_MATCH',
            'confidence': 0.0,
            'reasoning': 'Reference database is empty',
            'matched_description': None
        }

    # Extract and normalize the extracted product ID
    extracted_id = extracted_product.get('product_id', 'N/A')
    normalized_extracted = normalize_product_id(extracted_id)

    # If Product ID is missing or invalid, fallback to AI matching
    if not normalized_extracted or extracted_id == 'N/A':
        if use_ai_fallback:
            print(f"  ⚠️  Product ID missing/invalid ({extracted_id}), using AI matching with description...")
            return match_product_with_gemini(extracted_product, reference_database, confidence_threshold)
        else:
            return {
                'matched_id': 'NO_MATCH',
                'confidence': 0.0,
                'reasoning': f'Invalid extracted product ID: {extracted_id} (AI fallback disabled)',
                'matched_description': None
            }

    # Search for matching reference products
    # Try two strategies:
    # 1. Exact match: "11" matches "11"
    # 2. With -01 suffix: "11" matches "11-01" (for K002XXX products)
    matching_products = []
    for ref_product in reference_database:
        ref_id = ref_product['product_id']
        normalized_ref = normalize_product_id(ref_id)

        # Strategy 1: Exact match
        if normalized_extracted == normalized_ref:
            matching_products.append(ref_product)
        # Strategy 2: Try adding -01 suffix to extracted ID
        elif normalized_extracted + '-01' == normalized_ref:
            matching_products.append(ref_product)

    # No match found
    if len(matching_products) == 0:
        return {
            'matched_id': 'NO_MATCH',
            'confidence': 0.0,
            'reasoning': f'No match found for normalized ID: {normalized_extracted} (from {extracted_id})',
            'matched_description': None
        }

    # Single match found - return it
    if len(matching_products) == 1:
        matched = matching_products[0]
        return {
            'matched_id': matched['product_id'],
            'confidence': 1.0,
            'reasoning': f'ID match: {extracted_id} -> {matched["product_id"]} (normalized: {normalized_extracted})',
            'matched_description': matched['description']
        }

    # Multiple matches found - try to pick the simplest one first
    # Prefer IDs without extra suffixes (e.g., K002019-01 over K002019-01FZ)
    simplest_match = select_simplest_product_id(matching_products)

    if simplest_match:
        # Found a clear simplest match
        return {
            'matched_id': simplest_match['product_id'],
            'confidence': 1.0,
            'reasoning': f'ID match (simplest): {extracted_id} -> {simplest_match["product_id"]} (normalized: {normalized_extracted})',
            'matched_description': simplest_match['description']
        }

    # No clear simplest match - check if ID is short (ambiguous)
    is_short_id = normalized_extracted.isdigit() and len(normalized_extracted) <= 2

    if is_short_id and use_ai_fallback:
        # Short ID with multiple matches - use AI to disambiguate
        print(f"  ⚠️  Ambiguous short ID '{normalized_extracted}' matches {len(matching_products)} products, using AI to verify...")
        print(f"      Candidates: {', '.join([p['product_id'] for p in matching_products])}")

        # Use AI matching with only the matching candidates
        return match_product_with_gemini(extracted_product, matching_products, confidence_threshold)
    else:
        # Multiple matches but not short ID - return first match (shouldn't happen often)
        matched = matching_products[0]
        return {
            'matched_id': matched['product_id'],
            'confidence': 0.9,
            'reasoning': f'ID match (multiple candidates): {extracted_id} -> {matched["product_id"]} (normalized: {normalized_extracted})',
            'matched_description': matched['description']
        }


def match_product_with_gemini(
    extracted_product: Dict[str, str],
    reference_database: List[Dict[str, str]],
    confidence_threshold: float = 0.8
) -> Dict[str, any]:
    """
    Match an extracted product against reference database using Gemini AI.

    NOTE: This function considers BOTH Product ID and Description.
    For ID-only matching, use match_product_by_id() instead.

    Args:
        extracted_product: Dict with keys 'product_id' and 'description'
        reference_database: List of reference products from load_reference_database()
        confidence_threshold: Minimum confidence (0.0-1.0) to return a match

    Returns:
        Dictionary with keys:
        - matched_id: The matched reference product ID or "NO_MATCH"
        - confidence: Confidence score (0.0-1.0) or None
        - reasoning: Explanation of the match or None
        - matched_description: Description of matched product or None

    Raises:
        Exception: If Gemini API is not configured or request fails
    """
    if not GEMINI_API_KEY:
        raise Exception("GEMINI_API_KEY environment variable not set")

    # Handle empty reference database
    if not reference_database:
        return {
            'matched_id': 'NO_MATCH',
            'confidence': None,
            'reasoning': 'Reference database is empty',
            'matched_description': None
        }

    # Extract product info
    extracted_id = extracted_product.get('product_id', 'N/A')
    extracted_desc = extracted_product.get('description', 'N/A')

    # Build the prompt
    prompt = build_matching_prompt(extracted_id, extracted_desc, reference_database, confidence_threshold)

    try:
        # Use Gemini Flash for faster matching (this is a simpler task than PDF parsing)
        model = genai.GenerativeModel('gemini-2.0-flash-exp')

        response = model.generate_content(prompt)

        # Parse the response
        result = parse_gemini_matching_response(response.text, reference_database)

        return result

    except Exception as e:
        # Return error as NO_MATCH
        return {
            'matched_id': 'NO_MATCH',
            'confidence': None,
            'reasoning': f'Error calling Gemini API: {str(e)}',
            'matched_description': None
        }


def build_matching_prompt(
    extracted_id: str,
    extracted_desc: str,
    reference_database: List[Dict[str, str]],
    confidence_threshold: float
) -> str:
    """
    Build the Gemini prompt for product matching.
    
    Args:
        extracted_id: Extracted product ID
        extracted_desc: Extracted product description
        reference_database: List of reference products
        confidence_threshold: Minimum confidence threshold
    
    Returns:
        Formatted prompt string
    """
    # Limit reference database to first 500 products to avoid token limits
    # In production, you might want to pre-filter based on fuzzy string matching
    limited_db = reference_database[:500]
    
    # Format reference products as a numbered list
    reference_list = "\n".join([
        f"{i+1}. ID: {prod['product_id']} | Description: {prod['description']}"
        for i, prod in enumerate(limited_db)
    ])
    
    prompt = f"""You are a product matching expert. Your task is to match an extracted product against a reference database.

**EXTRACTED PRODUCT:**
- Product ID: {extracted_id}
- Description: {extracted_desc}

**REFERENCE DATABASE ({len(limited_db)} products):**
{reference_list}

**MATCHING INSTRUCTIONS:**
1. Consider BOTH the Product ID and Description when matching
2. **CRITICAL RULE**: If Product ID matches EXACTLY but the Description describes a COMPLETELY DIFFERENT product type (e.g., "CHICKEN BREAST" vs "BEEF BRISKET", or "PORK" vs "FISH"), you MUST return "NO_MATCH" with confidence 0.0. This indicates a data quality issue.
3. Handle variations in Product IDs:
   - Extra prefixes (e.g., "TYS-Z75880" should match "Z75880")
   - Extra suffixes (e.g., "Z75880-CS" should match "Z75880")
   - Different separators (e.g., "Z-75880" should match "Z75880")
4. Handle variations in Descriptions:
   - Different word order
   - Additional words (brand names, packaging info)
   - Abbreviations vs. full words
   - Case differences
5. **IMPORTANT**: Both Product ID AND Description must align. If they conflict, return "NO_MATCH"
6. Only return a match if you are at least {int(confidence_threshold * 100)}% confident
7. If no good match exists, return "NO_MATCH"

**OUTPUT FORMAT (JSON only, no other text):**
{{
    "matched_id": "the reference Product ID that matches, or NO_MATCH",
    "confidence": 0.95,
    "reasoning": "brief explanation of why this is a match"
}}

**IMPORTANT:**
- Return ONLY valid JSON, no markdown, no code blocks, no extra text
- The matched_id must be EXACTLY as it appears in the reference database
- Confidence must be a number between 0.0 and 1.0
- If no match, use: {{"matched_id": "NO_MATCH", "confidence": 0.0, "reasoning": "explanation"}}
"""
    
    return prompt


def parse_gemini_matching_response(
    response_text: str,
    reference_database: List[Dict[str, str]]
) -> Dict[str, any]:
    """
    Parse Gemini's matching response.
    
    Args:
        response_text: Raw response from Gemini
        reference_database: Reference database to look up matched description
    
    Returns:
        Dictionary with matched_id, confidence, reasoning, matched_description
    """
    try:
        # Clean up response (remove markdown code blocks if present)
        cleaned = response_text.strip()
        if cleaned.startswith('```'):
            # Remove markdown code blocks
            cleaned = re.sub(r'^```(?:json)?\s*\n', '', cleaned)
            cleaned = re.sub(r'\n```\s*$', '', cleaned)
        
        # Parse JSON
        result = json.loads(cleaned)
        
        matched_id = result.get('matched_id', 'NO_MATCH')
        confidence = result.get('confidence', 0.0)
        reasoning = result.get('reasoning', 'No reasoning provided')
        
        # Look up matched description
        matched_description = None
        if matched_id != 'NO_MATCH':
            for prod in reference_database:
                if prod['product_id'] == matched_id:
                    matched_description = prod['description']
                    break
        
        return {
            'matched_id': matched_id,
            'confidence': confidence,
            'reasoning': reasoning,
            'matched_description': matched_description
        }
        
    except json.JSONDecodeError as e:
        # Fallback: try to extract matched_id from text
        match = re.search(r'"matched_id"\s*:\s*"([^"]+)"', response_text)
        if match:
            matched_id = match.group(1)
            return {
                'matched_id': matched_id,
                'confidence': None,
                'reasoning': 'Parsed from malformed JSON',
                'matched_description': None
            }
        else:
            return {
                'matched_id': 'NO_MATCH',
                'confidence': None,
                'reasoning': f'Failed to parse Gemini response: {str(e)}',
                'matched_description': None
            }


def batch_match_products(
    extracted_products: List[Dict[str, str]],
    reference_database: List[Dict[str, str]],
    confidence_threshold: float = 0.8,
    verbose: bool = True,
    use_id_only: bool = True,
    use_ai_fallback: bool = True
) -> List[Dict[str, any]]:
    """
    Match multiple products in batch.

    Args:
        extracted_products: List of extracted products
        reference_database: Reference database
        confidence_threshold: Minimum confidence threshold (used for AI matching)
        verbose: Print progress messages
        use_id_only: If True, use ID-only matching (default). If False, use AI matching.
        use_ai_fallback: If True, fallback to AI when Product ID is missing (default: True)

    Returns:
        List of matching results (same order as input)
    """
    results = []

    for i, product in enumerate(extracted_products, 1):
        if verbose:
            print(f"Matching product {i}/{len(extracted_products)}...")

        if use_id_only:
            # Use fast ID-only matching (with AI fallback for missing IDs)
            result = match_product_by_id(
                product,
                reference_database,
                confidence_threshold=confidence_threshold,
                use_ai_fallback=use_ai_fallback
            )
        else:
            # Use AI matching (slower, considers descriptions)
            result = match_product_with_gemini(product, reference_database, confidence_threshold)

        results.append(result)

    return results


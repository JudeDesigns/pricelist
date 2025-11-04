from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, make_response
import os
from werkzeug.utils import secure_filename
import uuid
from pdf_processor import process_pdf, VENDOR_CONFIGS
import re
import json
import csv
from io import StringIO, BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
from datetime import datetime
from openpyxl import load_workbook

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()

# Configure tesseract path for Linux systems
try:
    import pytesseract
    # Try to find tesseract in common Linux locations
    tesseract_paths = ['/usr/bin/tesseract', '/usr/local/bin/tesseract', '/bin/tesseract']
    for path in tesseract_paths:
        if os.path.exists(path):
            pytesseract.pytesseract.tesseract_cmd = path
            break
except ImportError:
    pass  # pytesseract not installed, will be handled by pdf_processor

app = Flask(__name__)
# Use secret key from environment or fallback to default
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'your-secret-key-change-this-in-production')

# Debug: Print if Gemini API key is loaded
if os.environ.get('GEMINI_API_KEY'):
    print(f"✅ GEMINI_API_KEY loaded: {os.environ.get('GEMINI_API_KEY')[:20]}...")
else:
    print("⚠️  GEMINI_API_KEY not found in environment")

# Use absolute path for upload folder to avoid path issues
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size
app.config['ALLOWED_EXTENSIONS'] = {'pdf'}

# Create upload folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Vendor name mapping (filename patterns to vendor codes)
# NOTE: Order matters! More specific patterns should come first
VENDOR_NAME_MAPPING = {
    # Quirch Foods patterns
    'quirch foods': 'quirch_foods',
    'quirchfoods': 'quirch_foods',
    'quirch': 'quirch_foods',

    # Kruse & Sons patterns
    'kruse & sons': 'kruse_sons',
    'kruse and sons': 'kruse_sons',
    'kruse sons': 'kruse_sons',
    'kruse__son': 'kruse_sons',  # Handle underscores in filename
    'kruse': 'kruse_sons',

    # Glen Rose patterns
    'glen rose': 'glen_rose',
    'glenrose': 'glen_rose',
    'glen': 'glen_rose',

    # Purcell patterns
    'purcell': 'purcell',

    # Laras Meat patterns
    'laras meat': 'laras_meat',
    'laras_meat': 'laras_meat',
    'larasmeat': 'laras_meat',
    'laras': 'laras_meat',

    # Maui Prices patterns
    'maui prices': 'maui_prices',
    'maui_prices': 'maui_prices',
    'mauiprices': 'maui_prices',
    'maui': 'maui_prices',

    # C&D International Fishery patterns
    'c&d international': 'cd_international',
    'cd international': 'cd_international',
    'c&d': 'cd_international',
    'cd': 'cd_international',

    # Royalty Distribution patterns
    'royalty distribution': 'royalty_distribution',
    'royalty_distribution': 'royalty_distribution',
    'royaltydistribution': 'royalty_distribution',
    'royalty': 'royalty_distribution',

    # Los Angeles Poultry Co patterns
    'los_angeles_poultry': 'la_poultry',
    'los angeles poultry': 'la_poultry',
    'la poultry': 'la_poultry',
    'lapoultry': 'la_poultry',
    'la_poultry': 'la_poultry',
    'price_list_la': 'la_poultry',  # Price_List_LA filename pattern

    # TNT Produce Company patterns
    'tnt_produce_company': 'tnt_produce',
    'tnt produce company': 'tnt_produce',
    'tnt_produce': 'tnt_produce',
    'tnt produce': 'tnt_produce',
    'tntproduce': 'tnt_produce',
    'tnt': 'tnt_produce',

    # APSIC Wholesale patterns
    'apsic_wholesale': 'apsic_wholesale',
    'apsic wholesale': 'apsic_wholesale',
    'apsicwholesale': 'apsic_wholesale',
    'apsic': 'apsic_wholesale',

    # Del Mar Distributions COW patterns (order matters - most specific first)
    'del_mar_distributors_cow': 'delmar_cow',
    'del_mar_distributions_cow': 'delmar_cow',
    'del mar distributors cow': 'delmar_cow',
    'del mar distributions cow': 'delmar_cow',
    'delmar distributors cow': 'delmar_cow',
    'delmar distributions cow': 'delmar_cow',
    'del_mar_cow': 'delmar_cow',
    'delmar cow': 'delmar_cow',
    'delmarcow': 'delmar_cow',
    'del mar cow': 'delmar_cow',

    # Del Mar Distributions Steer patterns (order matters - most specific first)
    'del_mar_distributors_steer': 'delmar_steer',
    'del_mar_distributions_steer': 'delmar_steer',
    'del mar distributors steer': 'delmar_steer',
    'del mar distributions steer': 'delmar_steer',
    'delmar distributors steer': 'delmar_steer',
    'delmar distributions steer': 'delmar_steer',
    'del_mar_steer': 'delmar_steer',
    'delmar steer': 'delmar_steer',
    'delmarsteer': 'delmar_steer',
    'del mar steer': 'delmar_steer',

    # Gladway Pricing patterns
    'gladway_pricing': 'gladway',
    'gladway pricing': 'gladway',
    'gladwaypricing': 'gladway',
    'gladway': 'gladway',

    # Union Fish patterns
    'union_fish': 'union_fish',
    'union fish': 'union_fish',
    'unionfish': 'union_fish',

    # Solomon Wholesale patterns
    'solomon_wholesale_prices': 'solomon_wholesale',
    'solomon_wholesale': 'solomon_wholesale',
    'solomon wholesale prices': 'solomon_wholesale',
    'solomon wholesale': 'solomon_wholesale',
    'solomonwholesale': 'solomon_wholesale',
    'solomon': 'solomon_wholesale',

    # D&A PRICE patterns
    'd&a_price': 'da_price',
    'd&a price': 'da_price',
    'da_price': 'da_price',
    'da price': 'da_price',
    'daprice': 'da_price',
    'd&a': 'da_price',

    # Broadleaf patterns
    'broadleaf': 'broadleaf',
    'broad_leaf': 'broadleaf',
    'broad leaf': 'broadleaf',

    # Cofoods Inc patterns
    'cofoods_inc': 'cofoods',
    'cofoods inc': 'cofoods',
    'cofoodsinc': 'cofoods',
    'cofoods': 'cofoods',
    'co_foods': 'cofoods',
    'co foods': 'cofoods',

    # Monarch Trading patterns
    'monarch_trading': 'monarch_trading',
    'monarch trading': 'monarch_trading',
    'monarchtrading': 'monarch_trading',
    'monarch': 'monarch_trading',

    # RW Zant patterns
    'rw zant': 'rw_zant',
    'rwzant': 'rw_zant',
    'zant': 'rw_zant',
}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def restructure_duplicate_products(data):
    """
    Restructure products with duplicate IDs into single rows with ONLY case_price and pallet_price columns.
    Only creates these two specific columns for vendors with duplicate product IDs.
    """
    # Group by product ID
    product_groups = {}
    for item in data:
        pid = item.get('product_id', '')
        if pid not in product_groups:
            product_groups[pid] = []
        product_groups[pid].append(item)

    restructured = []
    for pid, items in product_groups.items():
        if len(items) == 1:
            # No duplicates for this product - keep original structure
            restructured.append({
                'product_id': pid,
                'description': items[0].get('description', ''),
                'cost': items[0].get('cost', '')
            })
        else:
            # Multiple entries - create case_price and pallet_price columns ONLY
            base_row = {
                'product_id': pid,
                'description': '',
                'case_price': '',
                'pallet_price': ''
            }

            # Extract case and pallet prices from descriptions
            for item in items:
                desc = item.get('description', '').upper()
                cost = item.get('cost', '')

                # Identify if this is a CASE or PALLET price
                if 'CASE' in desc or 'CS' in desc or '/CS' in desc:
                    base_row['case_price'] = cost
                    # Use case description as base (cleaned)
                    if not base_row['description']:
                        clean_desc = desc
                        for keyword in ['CASE', 'CS', '/CS', '-CS', '- CS']:
                            clean_desc = clean_desc.replace(keyword, '')
                        base_row['description'] = clean_desc.strip()
                elif 'PALLET' in desc or 'PLT' in desc or '/PLT' in desc:
                    base_row['pallet_price'] = cost
                    # Use pallet description as base if case wasn't found
                    if not base_row['description']:
                        clean_desc = desc
                        for keyword in ['PALLET', 'PLT', '/PLT', '-PLT', '- PLT']:
                            clean_desc = clean_desc.replace(keyword, '')
                        base_row['description'] = clean_desc.strip()
                else:
                    # If we can't identify, use first as case, second as pallet
                    if not base_row['case_price']:
                        base_row['case_price'] = cost
                        if not base_row['description']:
                            base_row['description'] = item.get('description', '')
                    elif not base_row['pallet_price']:
                        base_row['pallet_price'] = cost

            # If description is still empty, use the first item's description
            if not base_row['description'] and items:
                base_row['description'] = items[0].get('description', '')

            restructured.append(base_row)

    return restructured

def extract_vendor_name(filename):
    """Extract vendor name from filename (string part before numbers/date)"""
    # Remove .pdf extension
    name = filename.rsplit('.', 1)[0]
    # Extract the string part before numbers/date
    # Match letters, spaces, underscores, hyphens at the beginning
    match = re.match(r'^([a-zA-Z\s_-]+)', name)
    if match:
        vendor_name = match.group(1).strip('_- ')
        return vendor_name if vendor_name else filename.rsplit('.', 1)[0]
    return name

def detect_vendor_code(filename):
    """Detect vendor code from filename"""
    vendor_name = extract_vendor_name(filename).lower()

    print(f"Vendor Name: {vendor_name}")  # Debug output

    # Try to match against known vendor patterns
    # Check in order (more specific patterns first)
    for pattern, code in VENDOR_NAME_MAPPING.items():
        if pattern in vendor_name:
            print(f"Matched pattern '{pattern}' -> vendor code: {code}")  # Debug output
            return code

    # Return 'unknown' if no match - will be handled by Gemini
    print(f"No pattern matched, marking as unknown vendor")  # Debug output
    return 'unknown'

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    if 'files[]' not in request.files:
        flash('No files selected', 'error')
        return redirect(url_for('index'))

    files = request.files.getlist('files[]')

    if not files or files[0].filename == '':
        flash('No files selected', 'error')
        return redirect(url_for('index'))

    # Validate file count
    if len(files) > 30:
        flash('Maximum 30 files allowed', 'error')
        return redirect(url_for('index'))

    # Check if there's an existing session
    existing_session_id = session.get('session_id')

    if existing_session_id:
        # Reuse existing session
        session_id = existing_session_id
        print(f"📂 Reusing existing session: {session_id}")
    else:
        # Create a new session ID
        session_id = str(uuid.uuid4())
        print(f"📂 Creating new session: {session_id}")

    session_folder = os.path.join(app.config['UPLOAD_FOLDER'], session_id)
    os.makedirs(session_folder, exist_ok=True)

    uploaded_files = []

    for file in files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(session_folder, filename)
            file.save(filepath)
            uploaded_files.append(filename)
        else:
            flash(f'Invalid file: {file.filename}', 'warning')

    if not uploaded_files:
        flash('No valid PDF files uploaded', 'error')
        return redirect(url_for('index'))

    # Store session info
    session['session_id'] = session_id
    session['uploaded_files'] = uploaded_files

    return redirect(url_for('process'))

@app.route('/process')
def process():
    session_id = session.get('session_id')
    uploaded_files = session.get('uploaded_files')
    
    if not session_id or not uploaded_files:
        flash('No files to process', 'error')
        return redirect(url_for('index'))
    
    return render_template('process.html', file_count=len(uploaded_files))

def process_single_pdf(filename, session_folder):
    """
    Process a single PDF file. This function is called by ThreadPoolExecutor.

    Args:
        filename: Name of the PDF file
        session_folder: Path to the session folder containing the PDF

    Returns:
        Dictionary with processing results
    """
    filepath = os.path.join(session_folder, filename)
    vendor_name = extract_vendor_name(filename)
    vendor_code = detect_vendor_code(filename)

    print(f"\n{'='*60}")
    print(f"[PARALLEL] Processing: {filename}")
    print(f"Vendor Name: {vendor_name}")
    print(f"Vendor Code: {vendor_code} ({VENDOR_CONFIGS[vendor_code]['name']})")
    print(f"{'='*60}")

    try:
        start_time = time.time()
        extracted_data = process_pdf(filepath, vendor=vendor_code)
        elapsed_time = time.time() - start_time

        print(f"\n✅ [{filename}] Completed in {elapsed_time:.2f}s - Extracted {len(extracted_data)} products")

        return {
            'vendor_name': vendor_name,
            'vendor_code': vendor_code,
            'filename': filename,
            'status': 'success',
            'data': extracted_data,
            'error': None,
            'processing_time': elapsed_time
        }
    except Exception as e:
        print(f"\n❌ [{filename}] Failed: {str(e)}")
        return {
            'vendor_name': vendor_name,
            'vendor_code': vendor_code,
            'filename': filename,
            'status': 'error',
            'data': [],
            'error': str(e),
            'processing_time': 0
        }


@app.route('/results')
def results():
    session_id = session.get('session_id')
    uploaded_files = session.get('uploaded_files')

    if not session_id or not uploaded_files:
        flash('No files to process', 'error')
        return redirect(url_for('index'))

    # Start processing in background thread
    session_folder = os.path.join(app.config['UPLOAD_FOLDER'], session_id)
    results_file = os.path.join(session_folder, 'results.json')

    # Check if there are new files to process
    existing_filenames = set()
    initial_results = []

    if os.path.exists(results_file):
        # Load existing results
        try:
            with open(results_file, 'r') as f:
                initial_results = json.load(f)
                existing_filenames = {v['filename'] for v in initial_results}
        except Exception as e:
            print(f"⚠️  Could not load existing results: {e}")
            initial_results = []

    # Check if there are new files to process
    new_files = [f for f in uploaded_files if f not in existing_filenames]

    if new_files:
        # There are new files - start background processing
        print(f"🆕 Found {len(new_files)} new files to process")
        import threading
        thread = threading.Thread(target=process_pdfs_background, args=(session_id, uploaded_files, session_folder))
        thread.daemon = True
        thread.start()
    else:
        print(f"✓ All {len(uploaded_files)} files already processed")

    # Always render the template (with existing results if any)
    return render_template('results_dynamic.html',
                         session_id=session_id,
                         total_files=len(uploaded_files),
                         initial_results=initial_results)

def process_pdfs_background(session_id, uploaded_files, session_folder):
    """Process PDFs in background and save results incrementally"""

    # 🚀 PARALLEL PROCESSING: Process multiple PDFs simultaneously
    print(f"\n{'='*80}")
    print(f"🚀 PARALLEL PROCESSING MODE (Background)")
    print(f"   Session ID: {session_id}")
    print(f"   Total files: {len(uploaded_files)}")
    print(f"   Max workers: 10 (processing 10 PDFs at once)")
    print(f"{'='*80}\n")

    start_time = time.time()
    results_file = os.path.join(session_folder, 'results.json')
    status_file = os.path.join(session_folder, 'status.json')

    # Load existing results if they exist
    vendor_results = []
    existing_filenames = set()
    if os.path.exists(results_file):
        try:
            with open(results_file, 'r') as f:
                vendor_results = json.load(f)
                existing_filenames = {v['filename'] for v in vendor_results}
            print(f"📂 Loaded {len(vendor_results)} existing results")
        except Exception as e:
            print(f"⚠️  Could not load existing results: {e}")
            vendor_results = []

    # Filter out files that have already been processed
    files_to_process = [f for f in uploaded_files if f not in existing_filenames]

    if not files_to_process:
        print(f"✓ All files already processed")
        # Update status to complete
        with open(status_file, 'w') as f:
            json.dump({
                'status': 'complete',
                'completed': len(vendor_results),
                'total': len(vendor_results),
                'total_time': 0,
                'total_products': sum(len(v.get('data', [])) for v in vendor_results if v['status'] == 'success')
            }, f)
        return

    print(f"📋 Processing {len(files_to_process)} new files (skipping {len(existing_filenames)} already processed)")

    # Initialize status
    with open(status_file, 'w') as f:
        json.dump({
            'status': 'processing',
            'completed': len(vendor_results),
            'total': len(vendor_results) + len(files_to_process)
        }, f)

    # Use ThreadPoolExecutor to process PDFs in parallel
    # max_workers=10 means process up to 10 PDFs simultaneously
    with ThreadPoolExecutor(max_workers=10) as executor:
        # Submit only new PDF processing tasks
        future_to_filename = {
            executor.submit(process_single_pdf, filename, session_folder): filename
            for filename in files_to_process
        }

        # Collect results as they complete
        for future in as_completed(future_to_filename):
            filename = future_to_filename[future]
            try:
                result = future.result()
                vendor_results.append(result)
                print(f"✓ Collected result for: {filename}")

                # Save incremental results
                with open(results_file, 'w') as f:
                    json.dump(vendor_results, f)

                # Update status
                with open(status_file, 'w') as f:
                    json.dump({
                        'status': 'processing',
                        'completed': len(vendor_results),
                        'total': len(vendor_results) + len(files_to_process) - len([r for r in vendor_results if r['filename'] in files_to_process])
                    }, f)

            except Exception as e:
                print(f"✗ Error collecting result for {filename}: {str(e)}")
                vendor_results.append({
                    'vendor_name': extract_vendor_name(filename),
                    'vendor_code': detect_vendor_code(filename),
                    'filename': filename,
                    'status': 'error',
                    'data': [],
                    'error': f"Processing failed: {str(e)}",
                    'processing_time': 0
                })

                # Save incremental results
                with open(results_file, 'w') as f:
                    json.dump(vendor_results, f)

    total_time = time.time() - start_time
    total_products = sum(len(v['data']) for v in vendor_results if v['status'] == 'success')

    print(f"\n{'='*80}")
    print(f"🎉 PARALLEL PROCESSING COMPLETE!")
    print(f"   Total time: {total_time:.2f}s")
    print(f"   Files processed: {len(vendor_results)} total ({len(files_to_process)} new)")
    print(f"   Total products extracted: {total_products}")
    if files_to_process:
        print(f"   Average time per file: {total_time/len(files_to_process):.2f}s")
    print(f"{'='*80}\n")

    # Update final status
    with open(status_file, 'w') as f:
        json.dump({
            'status': 'complete',
            'completed': len(vendor_results),
            'total': len(vendor_results),
            'total_time': total_time,
            'total_products': total_products
        }, f)

@app.route('/api/results/<session_id>')
def api_results(session_id):
    """API endpoint for polling results"""
    from flask import jsonify

    session_folder = os.path.join(app.config['UPLOAD_FOLDER'], session_id)
    results_file = os.path.join(session_folder, 'results.json')
    status_file = os.path.join(session_folder, 'status.json')

    # Load status
    status_data = {'status': 'processing', 'completed': 0, 'total': 0}
    if os.path.exists(status_file):
        with open(status_file, 'r') as f:
            status_data = json.load(f)

    # Load results
    vendor_results = []
    if os.path.exists(results_file):
        with open(results_file, 'r') as f:
            vendor_results = json.load(f)

    return jsonify({
        'status': status_data.get('status', 'processing'),
        'completed': status_data.get('completed', 0),
        'total_files': status_data.get('total', 0),
        'results': vendor_results,
        'total_time': status_data.get('total_time'),
        'total_products': status_data.get('total_products')
    })

@app.route('/export/all/<format>')
def export_all_products(format):
    """Export all products from all vendors"""
    session_id = session.get('session_id')

    if not session_id:
        return "No data to export. Please process files first.", 400

    # Load results from file
    session_folder = os.path.join(app.config['UPLOAD_FOLDER'], session_id)
    results_file = os.path.join(session_folder, 'results.json')

    if not os.path.exists(results_file):
        return "No data to export. Please process files first.", 400

    with open(results_file, 'r') as f:
        vendor_results = json.load(f)

    # Combine all products
    all_products = []
    for vendor in vendor_results:
        if vendor['status'] == 'success' and vendor.get('data'):
            for product in vendor['data']:
                all_products.append({
                    'vendor_name': vendor['vendor_name'],
                    'vendor_code': vendor['vendor_code'],
                    'filename': vendor['filename'],
                    'product_id': product.get('product_id', 'N/A'),
                    'product_description': product.get('description', product.get('product_description', '')),
                    'cost': product.get('cost', 0)
                })

    if not all_products:
        return "No data to export.", 400

    if format == 'csv':
        # Create CSV
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=['vendor_name', 'vendor_code', 'filename', 'product_id', 'product_description', 'cost'])
        writer.writeheader()
        writer.writerows(all_products)

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=all_products_export.csv'
        return response

    elif format == 'json':
        # Create JSON
        response = make_response(json.dumps(all_products, indent=2))
        response.headers['Content-Type'] = 'application/json'
        response.headers['Content-Disposition'] = f'attachment; filename=all_products_export.json'
        return response

    return "Invalid format", 400

@app.route('/export/<int:vendor_index>/<format>')
def export_data(vendor_index, format):
    """Export vendor data in CSV or JSON format"""
    session_id = session.get('session_id')

    if not session_id:
        return "No data to export. Please process files first.", 400

    # Load results from file
    session_folder = os.path.join(app.config['UPLOAD_FOLDER'], session_id)
    results_file = os.path.join(session_folder, 'results.json')

    if not os.path.exists(results_file):
        return "Results file not found. Please process files again.", 400

    try:
        with open(results_file, 'r') as f:
            vendor_results = json.load(f)
    except Exception as e:
        return f"Error loading results: {str(e)}", 500

    # Debug logging
    print(f"\n{'='*60}")
    print(f"EXPORT REQUEST:")
    print(f"  Vendor Index: {vendor_index}")
    print(f"  Format: {format}")
    print(f"  Session ID: {session_id}")
    print(f"  Results loaded from file: {results_file}")
    if vendor_results:
        print(f"  Total vendors: {len(vendor_results)}")
        for i, v in enumerate(vendor_results):
            print(f"    [{i}] {v.get('filename')} - Status: {v.get('status')} - Items: {len(v.get('data', []))}")
    print(f"{'='*60}\n")

    if not vendor_results:
        return "No data to export. Please process files first.", 400

    if vendor_index >= len(vendor_results):
        return f"Invalid vendor index: {vendor_index}. Only {len(vendor_results)} vendors available.", 400

    vendor = vendor_results[vendor_index]

    if vendor['status'] != 'success' or not vendor['data']:
        return f"No data available for vendor: {vendor.get('vendor_name', 'Unknown')} (Status: {vendor['status']}, Items: {len(vendor.get('data', []))})", 400

    # Use the original filename (without extension) for the export
    original_filename = vendor.get('filename', 'export')
    # Remove .pdf extension if present
    if original_filename.lower().endswith('.pdf'):
        original_filename = original_filename[:-4]

    # Sanitize filename - keep alphanumeric, spaces, hyphens, underscores
    safe_filename = re.sub(r'[^\w\s-]', '', original_filename).strip().replace(' ', '_')

    if format == 'csv':
        # Check for duplicate product IDs
        product_id_counts = {}
        for item in vendor['data']:
            pid = item.get('product_id', '')
            product_id_counts[pid] = product_id_counts.get(pid, 0) + 1

        has_duplicates = any(count > 1 for count in product_id_counts.values())

        if has_duplicates:
            # Handle duplicates - create columns for each price type
            print(f"⚠️  Duplicate product IDs detected - restructuring export")
            restructured_data = restructure_duplicate_products(vendor['data'])

            # Determine fieldnames dynamically based on price columns
            fieldnames = ['product_id', 'description']
            price_columns = set()
            for item in restructured_data:
                for key in item.keys():
                    if key not in ['product_id', 'description']:
                        price_columns.add(key)
            fieldnames.extend(sorted(price_columns))

            output = StringIO()
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(restructured_data)
        else:
            # No duplicates - standard export
            output = StringIO()
            writer = csv.DictWriter(output, fieldnames=['product_id', 'description', 'cost'])
            writer.writeheader()

            for item in vendor['data']:
                writer.writerow({
                    'product_id': item.get('product_id', ''),
                    'description': item.get('description', ''),
                    'cost': item.get('cost', '')
                })

        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={safe_filename}.csv'
        response.headers['Content-Type'] = 'text/csv'
        return response

    elif format == 'json':
        # Generate JSON
        json_data = json.dumps(vendor['data'], indent=2)

        response = make_response(json_data)
        response.headers['Content-Disposition'] = f'attachment; filename={safe_filename}.json'
        response.headers['Content-Type'] = 'application/json'
        return response

    else:
        return f"Invalid export format: {format}. Use 'csv' or 'json'.", 400

@app.route('/update-spreadsheet', methods=['POST'])
def update_spreadsheet():
    """Update an existing Excel spreadsheet with new prices from extracted data"""
    session_id = session.get('session_id')

    if not session_id:
        return "No data available. Please process files first.", 400

    # Load results from file
    session_folder = os.path.join(app.config['UPLOAD_FOLDER'], session_id)
    results_file = os.path.join(session_folder, 'results.json')

    if not os.path.exists(results_file):
        return "No data to export. Please process files first.", 400

    with open(results_file, 'r') as f:
        vendor_results = json.load(f)

    # Get uploaded spreadsheet
    if 'spreadsheet' not in request.files:
        return "No spreadsheet file uploaded.", 400

    spreadsheet_file = request.files['spreadsheet']

    if spreadsheet_file.filename == '':
        return "No spreadsheet file selected.", 400

    if not spreadsheet_file.filename.endswith('.xlsx'):
        return "Invalid file format. Please upload an Excel (.xlsx) file.", 400

    # Get form parameters
    price_type = request.form.get('price_type', 'cost')
    product_id_column = request.form.get('product_id_column', 'Product Code')

    try:
        # Load the workbook
        wb = load_workbook(spreadsheet_file)
        ws = wb.active

        # Find column indices
        header_row = None
        product_id_col_idx = None
        current_cost_col_idx = None
        old_cost_col_idx = None
        last_updated_col_idx = None

        # Search for headers in first 10 rows
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row = ws[row_idx]
            for cell_idx, cell in enumerate(row, start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()

                    # Check for Product ID column
                    if cell_value.lower() in [product_id_column.lower(), 'product code', 'product id', 'sku', 'item code']:
                        product_id_col_idx = cell_idx
                        header_row = row_idx

                    # Check for Current Cost column
                    if cell_value.lower() in ['current cost', 'cost', 'price', 'unit price']:
                        current_cost_col_idx = cell_idx
                        header_row = row_idx

                    # Check for Old Cost column
                    if cell_value.lower() in ['old cost', 'previous cost', 'last cost']:
                        old_cost_col_idx = cell_idx

                    # Check for Last Updated column
                    if cell_value.lower() in ['last updated', 'date updated', 'updated']:
                        last_updated_col_idx = cell_idx

            # If we found the key columns, stop searching
            if product_id_col_idx and current_cost_col_idx:
                break

        if not product_id_col_idx:
            return f"Could not find '{product_id_column}' column in spreadsheet. Please check the column name.", 400

        if not current_cost_col_idx:
            return "Could not find 'Current Cost' column in spreadsheet.", 400

        # Create "Old Cost" column if it doesn't exist
        if not old_cost_col_idx:
            old_cost_col_idx = ws.max_column + 1
            ws.cell(row=header_row, column=old_cost_col_idx, value="Old Cost")

        # Create "Last Updated" column if it doesn't exist
        if not last_updated_col_idx:
            last_updated_col_idx = ws.max_column + 1
            ws.cell(row=header_row, column=last_updated_col_idx, value="Last Updated")

        # Build a lookup dictionary from extracted data
        price_lookup = {}
        for vendor in vendor_results:
            if vendor['status'] == 'success' and vendor.get('data'):
                for product in vendor['data']:
                    product_id = str(product.get('product_id', '')).strip()

                    if product_id and product_id != 'N/A':
                        # Determine which price to use
                        if price_type == 'case_price' and 'case_price' in product:
                            price = product.get('case_price', '')
                        elif price_type == 'pallet_price' and 'pallet_price' in product:
                            price = product.get('pallet_price', '')
                        else:
                            price = product.get('cost', '')

                        if price:
                            # Clean price (remove $ if present)
                            price_str = str(price).replace('$', '').strip()
                            price_lookup[product_id.upper()] = price_str

        # Update the spreadsheet
        updated_count = 0
        today = datetime.now().strftime('%Y-%m-%d')

        for row_idx in range(header_row + 1, ws.max_row + 1):
            product_id_cell = ws.cell(row=row_idx, column=product_id_col_idx)
            product_id = str(product_id_cell.value).strip() if product_id_cell.value else ''

            if product_id and product_id.upper() in price_lookup:
                # Get current cost
                current_cost_cell = ws.cell(row=row_idx, column=current_cost_col_idx)
                current_cost = current_cost_cell.value

                # Move current cost to old cost
                old_cost_cell = ws.cell(row=row_idx, column=old_cost_col_idx)
                old_cost_cell.value = current_cost

                # Update current cost with new price
                new_price = price_lookup[product_id.upper()]
                current_cost_cell.value = new_price

                # Update last updated date
                last_updated_cell = ws.cell(row=row_idx, column=last_updated_col_idx)
                last_updated_cell.value = today

                updated_count += 1

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        original_filename = secure_filename(spreadsheet_file.filename)
        if original_filename.lower().endswith('.xlsx'):
            original_filename = original_filename[:-5]

        new_filename = f"Updated_Prices_{original_filename}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{new_filename}"'

        print(f"✅ Updated {updated_count} products in spreadsheet")

        return response

    except Exception as e:
        print(f"❌ Error updating spreadsheet: {str(e)}")
        import traceback
        traceback.print_exc()
        return f"Error processing spreadsheet: {str(e)}", 500

@app.route('/ai-match', methods=['POST'])
def ai_match():
    """AI-powered product matching endpoint"""
    try:
        from ai_matcher import load_reference_database, batch_match_products

        # Get confidence threshold
        confidence_threshold = float(request.form.get('confidence_threshold', 0.8))

        # Check session
        session_id = session.get('session_id')
        if not session_id:
            return "No active session. Please upload PDFs first.", 400

        # Use Book1.xlsx from project root (hardcoded reference database)
        reference_path = os.path.join(os.path.dirname(__file__), 'Book1.xlsx')

        if not os.path.exists(reference_path):
            return f"Reference database not found at {reference_path}", 500

        # Load reference database
        print(f"Loading reference database from {reference_path}...")
        reference_db = load_reference_database(reference_path)
        print(f"Loaded {len(reference_db)} reference products")

        # Get extracted products from results file
        session_folder = os.path.join(app.config['UPLOAD_FOLDER'], session_id)
        results_file = os.path.join(session_folder, 'results.json')

        if not os.path.exists(results_file):
            return "No results found. Please upload and process PDFs first.", 400

        with open(results_file, 'r') as f:
            vendor_results = json.load(f)

        # Combine all products from all vendors
        all_products = []
        for vendor in vendor_results:
            if vendor['status'] == 'success' and vendor.get('data'):
                for product in vendor['data']:
                    all_products.append({
                        'product_id': product.get('product_id', 'N/A'),
                        'description': product.get('description', product.get('product_description', '')),
                        'vendor': vendor['vendor_name']
                    })

        if not all_products:
            return "No products found in results. Please process PDFs first.", 400

        print(f"Found {len(all_products)} extracted products to match")

        # Prepare products for matching
        extracted_products = all_products

        print(f"Matching {len(extracted_products)} extracted products...")

        # Perform ID-only matching (fast, no AI needed)
        match_results = batch_match_products(
            extracted_products,
            reference_db,
            confidence_threshold=confidence_threshold,
            verbose=True,
            use_id_only=True  # Use ID-only matching (strips leading Z and trailing letters)
        )

        # Prepare response
        results = []
        matched_count = 0
        total_confidence = 0
        confidence_count = 0

        for extracted, match_result in zip(extracted_products, match_results):
            result = {
                'extracted_id': extracted['product_id'],
                'extracted_description': extracted['description'],
                'matched_id': match_result['matched_id'],
                'matched_description': match_result.get('matched_description', ''),
                'confidence': match_result.get('confidence', 0),
                'reasoning': match_result.get('reasoning', '')
            }
            results.append(result)

            if match_result['matched_id'] != 'NO_MATCH':
                matched_count += 1
                if match_result.get('confidence'):
                    total_confidence += match_result['confidence']
                    confidence_count += 1

        # Calculate summary
        avg_confidence = (total_confidence / confidence_count * 100) if confidence_count > 0 else 0

        summary = {
            'total': len(extracted_products),
            'matched': matched_count,
            'unmatched': len(extracted_products) - matched_count,
            'avg_confidence': f"{avg_confidence:.1f}%"
        }

        return jsonify({
            'success': True,
            'summary': summary,
            'results': results
        })

    except Exception as e:
        print(f"Error in AI matching: {str(e)}")
        import traceback
        traceback.print_exc()
        return f"Error during AI matching: {str(e)}", 500

@app.route('/clear')
def clear_session():
    """Clear the current session and uploaded files"""
    session_id = session.get('session_id')
    if session_id:
        session_folder = os.path.join(app.config['UPLOAD_FOLDER'], session_id)
        # Clean up uploaded files
        if os.path.exists(session_folder):
            import shutil
            shutil.rmtree(session_folder)

    session.clear()
    flash('Session cleared', 'success')
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True, port=5001)

